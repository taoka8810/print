import openpyxl as px, random, shutil


#対象となるwordsファイルの指定と読み込み
fileNum=int(input(
"""
--------------------------------------------
<生徒ナンバーを入力してください>

1.中1　〇〇さん

2.中2　☐☐さん

3.
--------------------------------------------
"""))

if fileNum==1:
    filePath=r"C:\Users\taoka\Desktop\program\python\main\print\words_G1K.xlsx"
elif fileNum==2:
    filePath=r"C:\Users\taoka\Desktop\program\python\main\print\words_G2M.xlsx"
else:
    print("ファイルが存在しません")
wb_words=px.load_workbook(filePath)


#Unitの指定と読み込み
print("--------------------------------------------\n")
for index, name in enumerate(wb_words.sheetnames):
    print(str(index)+". "+name+"\n")
unitNum=int(input("""
Unitを選択してください：

--------------------------------------------
"""))
ws_words=wb_words.worksheets[unitNum]
unitName=ws_words.title


#乱数生成
def generateNumber(n1, n2, n3):  #どこから、どこまでで、何個生成するか
      if n2+1 < n3:
          n3=n2
      aa=0
      x=[]
      while aa < n3:
          n=random.randint(n1, n2)
          if n in x:
              continue
          else:
              x.append(n)
              aa+=1
      return x

#英単語と意味の読み込み
N=ws_words.max_row
i=1
word_list=[]
mean_list=[]
while i<=N:
    word_list.append(ws_words.cell(row=i, column=1).value)
    mean_list.append(ws_words.cell(row=i, column=2).value)
    i+=1

#sheet.xlsxをDesktopに複製
path=r"C:\Users\taoka\Desktop\program\python\main\print\sheet.xlsx"
desktop=r"C:\Users\taoka\Desktop\English.xlsx"
shutil.copy(path,desktop)

#English.xlsxの読み込み
wb_sheet=px.load_workbook(r"C:\Users\taoka\Desktop\English.xlsx")
s1=wb_sheet["1"]
s2=wb_sheet["2"]
s3=wb_sheet["3"]
s4=wb_sheet["4"]
s5=wb_sheet["5"]
s6=wb_sheet["6"]
s7=wb_sheet["7"]

#English.xlsxに単語と意味の書き込み
a=5
b=22
k=generateNumber(0, N-1, 35)
for j in k:
    if a<=21:
        word=word_list[j]
        mean=mean_list[j]
        s1.cell(row=a, column=1).value=word
        s1.cell(row=a, column=10).value=word
        s1.cell(row=a, column=14).value=mean
        a+=1
    elif b<=39:
        word=word_list[j]
        mean=mean_list[j]
        s1.cell(row=b, column=1).value=mean
        s1.cell(row=b, column=10).value=mean
        s1.cell(row=b, column=14).value=word
        b+=1

#2回目
a=5
b=22
k=generateNumber(0, N-1, 35)
for j in k:
    if a<=21:
        word=word_list[j]
        mean=mean_list[j]
        s2.cell(row=a, column=1).value=word
        s2.cell(row=a, column=10).value=word
        s2.cell(row=a, column=14).value=mean
        a+=1
    elif b<=39:
        word=word_list[j]
        mean=mean_list[j]
        s2.cell(row=b, column=1).value=mean
        s2.cell(row=b, column=10).value=mean
        s2.cell(row=b, column=14).value=word
        b+=1

#3回目
a=5
b=22
k=generateNumber(0, N-1, 35)
for j in k:
    if a<=21:
        word=word_list[j]
        mean=mean_list[j]
        s3.cell(row=a, column=1).value=word
        s3.cell(row=a, column=10).value=word
        s3.cell(row=a, column=14).value=mean
        a+=1
    elif b<=39:
        word=word_list[j]
        mean=mean_list[j]
        s3.cell(row=b, column=1).value=mean
        s3.cell(row=b, column=10).value=mean
        s3.cell(row=b, column=14).value=word
        b+=1

#4回目
a=5
b=22
k=generateNumber(0, N-1, 35)
for j in k:
    if a<=21:
        word=word_list[j]
        mean=mean_list[j]
        s4.cell(row=a, column=1).value=word
        s4.cell(row=a, column=10).value=word
        s4.cell(row=a, column=14).value=mean
        a+=1
    elif b<=39:
        word=word_list[j]
        mean=mean_list[j]
        s4.cell(row=b, column=1).value=mean
        s4.cell(row=b, column=10).value=mean
        s4.cell(row=b, column=14).value=word
        b+=1

#5回目
a=5
b=22
k=generateNumber(0, N-1, 35)
for j in k:
    if a<=21:
        word=word_list[j]
        mean=mean_list[j]
        s5.cell(row=a, column=1).value=word
        s5.cell(row=a, column=10).value=word
        s5.cell(row=a, column=14).value=mean
        a+=1
    elif b<=39:
        word=word_list[j]
        mean=mean_list[j]
        s5.cell(row=b, column=1).value=mean
        s5.cell(row=b, column=10).value=mean
        s5.cell(row=b, column=14).value=word
        b+=1

#6回目
a=5
b=22
k=generateNumber(0, N-1, 35)
for j in k:
    if a<=21:
        word=word_list[j]
        mean=mean_list[j]
        s6.cell(row=a, column=1).value=word
        s6.cell(row=a, column=10).value=word
        s6.cell(row=a, column=14).value=mean
        a+=1
    elif b<=39:
        word=word_list[j]
        mean=mean_list[j]
        s6.cell(row=b, column=1).value=mean
        s6.cell(row=b, column=10).value=mean
        s6.cell(row=b, column=14).value=word
        b+=1

#7回目
a=5
b=22
k=generateNumber(0, N-1, 35)
for j in k:
    if a<=21:
        word=word_list[j]
        mean=mean_list[j]
        s7.cell(row=a, column=1).value=word
        s7.cell(row=a, column=10).value=word
        s7.cell(row=a, column=14).value=mean
        a+=1
    elif b<=39:
        word=word_list[j]
        mean=mean_list[j]
        s7.cell(row=b, column=1).value=mean
        s7.cell(row=b, column=10).value=mean
        s7.cell(row=b, column=14).value=word
        b+=1

#タイトルの書き込み
font=px.styles.Font(size=14, bold=True)
title=unitName+"  単語プリント"
s1["A1"].value=title
s1["A1"].font=font
s2["A1"].value=title
s2["A1"].font=font
s3["A1"].value=title
s3["A1"].font=font
s4["A1"].value=title
s4["A1"].font=font
s5["A1"].value=title
s5["A1"].font=font
s6["A1"].value=title
s6["A1"].font=font
s7["A1"].value=title
s7["A1"].font=font

wb_sheet.save(r"C:\Users\taoka\Desktop\English.xlsx")
wb_sheet.close()
wb_words.close()

print("すべての処理が完了しました")
